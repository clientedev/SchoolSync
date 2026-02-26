import resend
import os
import uuid
from werkzeug.utils import secure_filename
from flask import current_app
from flask_mail import Message
from production_app import mail
from io import BytesIO

def send_resend_email(to_email, subject, html_content, attachments=None):
    """Send email using Resend API with optional attachments"""
    resend_api_key = os.environ.get("RESEND_API_KEY")
    if not resend_api_key:
        current_app.logger.warning("RESEND_API_KEY not configured - skipping email")
        return False
    
    try:
        resend.api_key = resend_api_key
        params = {
            "from": current_app.config.get('MAIL_DEFAULT_SENDER', 'onboarding@resend.dev'),
            "to": [to_email],
            "subject": subject,
            "html": html_content,
        }
        
        if attachments:
            params["attachments"] = attachments
            
        r = resend.Emails.send(params)
        current_app.logger.info(f"Email sent via Resend to {to_email}: {r}")
        return True
    except Exception as e:
        current_app.logger.error(f"Resend error sending email to {to_email}: {str(e)}")
        return False

def send_message_notification_resend(to_email, recipient_name, message_content):
    """Send notification email for a message"""
    subject = "Nova Mensagem do Sistema de Avaliação"
    html_content = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h1>Olá {recipient_name},</h1>
        <p>Você recebeu uma nova mensagem:</p>
        <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
            {message_content}
        </div>
        <p>Acesse o sistema para responder.</p>
    </div>
    """
    return send_resend_email(to_email, subject, html_content)

def send_signature_notification_evaluator(evaluation):
    """Send notification email to the evaluator when a teacher signs the evaluation"""
    if not evaluation.evaluator or not evaluation.evaluator.email:
        return False
    
    evaluator_email = evaluation.evaluator.email
    subject = f"Avaliação Assinada - {evaluation.teacher.name}"
    
    html_content = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h1 style="color: #2e7d32;">Avaliação Assinada</h1>
        <p>Olá <strong>{evaluation.evaluator.name}</strong>,</p>
        <p>O docente <strong>{evaluation.teacher.name}</strong> acabou de assinar digitalmente a avaliação realizada em {evaluation.evaluation_date.strftime('%d/%m/%Y')}.</p>
        
        <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <h2 style="margin-top: 0; font-size: 18px;">Resumo:</h2>
            <ul style="list-style: none; padding-left: 0;">
                <li><strong>Docente:</strong> {evaluation.teacher.name}</li>
                <li><strong>Data da Assinatura:</strong> {evaluation.teacher_signature_date.strftime('%d/%m/%Y %H:%M') if evaluation.teacher_signature_date else 'Agora'}</li>
                <li><strong>Curso:</strong> {evaluation.course.name if evaluation.course else 'N/A'}</li>
            </ul>
        </div>

        <p>A avaliação agora está formalmente concluída com ambas as assinaturas.</p>

        <p style="color: #666; font-size: 12px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 10px;">
            Este é um e-mail automático enviado pelo Sistema de Avaliação Docente SENAI Morvan Figueiredo.
        </p>
    </div>
    """
    return send_resend_email(evaluator_email, subject, html_content)

def send_evaluation_notification_resend(evaluation):
    """Send notification email when an evaluation is completed"""
    if not evaluation.teacher or not evaluation.teacher.user or not evaluation.teacher.user.email:
        return False
    
    teacher_email = evaluation.teacher.user.email
    teacher_user = evaluation.teacher.user
    subject = f"Nova Avaliação Docente - Assinatura Necessária - {evaluation.evaluation_date.strftime('%d/%m/%Y')}"
    
    # Credenciais e Link
    system_link = os.environ.get("DOMAIN") or "https://web-production-a43ed.up.railway.app/"
    username = teacher_user.username
    
    # Senha - Tenta pegar a senha real de forma mais abrangente
    password_info = None
    
    # PRIORIDADE: Seguir a mesma lógica do perfil para consistência total
    # 1. Tenta o campo permanentemente armazenado no banco
    if hasattr(teacher_user, 'plain_password') and teacher_user.plain_password:
        password_info = teacher_user.plain_password
    
    # 2. Tenta atributos temporários no objeto user (Request atual)
    if not password_info:
        if hasattr(teacher_user, '_password_plain') and teacher_user._password_plain:
            password_info = teacher_user._password_plain
        elif hasattr(teacher_user, '_temp_password') and teacher_user._temp_password:
            password_info = teacher_user._temp_password
    
    # 3. Tenta buscar na sessão (Última ação do usuário)
    if not password_info:
        from flask import session
        recent_creds = session.get('new_teacher_credentials')
        if recent_creds and recent_creds.get('username') == teacher_user.username:
            password_info = recent_creds.get('password')

    # 4. Fallback final (Caso nada funcione)
    if not password_info:
        password_info = teacher_user.username # Login/NIF como padrão de sistema

    # Gerar PDF de credenciais
    attachments = []
    try:
        from base64 import b64encode
        pdf_buffer = generate_teacher_credentials_pdf(
            evaluation.teacher.name, 
            evaluation.teacher.nif, 
            username, 
            password_info
        )
        # Resend expects base64 content for attachments
        import base64
        attachments.append({
            "filename": f"Credenciais_{username}.pdf",
            "content": base64.b64encode(pdf_buffer.getvalue()).decode('utf-8')
        })
    except Exception as pdf_err:
        current_app.logger.error(f"Erro ao gerar PDF de credenciais para anexo: {pdf_err}")

    html_content = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h1 style="color: #1976d2;">Finalização de Avaliação - Assinatura Necessária</h1>
        <p>Olá <strong>{evaluation.teacher.name}</strong>,</p>
        <p>Uma nova avaliação foi concluída para você e <strong>necessita de sua assinatura digital</strong>.</p>
        
        <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <h2 style="margin-top: 0; font-size: 18px;">Detalhes da Avaliação:</h2>
            <ul style="list-style: none; padding-left: 0;">
                <li><strong>Data:</strong> {evaluation.evaluation_date.strftime('%d/%m/%Y')}</li>
                <li><strong>Curso:</strong> {evaluation.course.name if evaluation.course else 'N/A'}</li>
                <li><strong>Avaliador:</strong> {evaluation.evaluator.name if evaluation.evaluator else 'Coordenação'}</li>
            </ul>
        </div>

        <div style="background-color: #e3f2fd; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <h2 style="margin-top: 0; font-size: 18px; color: #1976d2;">Acesso ao Sistema:</h2>
            <p>Suas credenciais de acesso (Login e Senha) estão no <strong>PDF em anexo</strong> a este e-mail.</p>
            <p style="margin-bottom: 5px;"><strong>Link do Sistema:</strong> <a href="{system_link}">{system_link}</a></p>
        </div>

        <p><strong>Próximos Passos:</strong></p>
        <ol>
            <li>Acesse o sistema pelo link acima.</li>
            <li>Entre com seu usuário e senha.</li>
            <li>Localize a avaliação na sua área de docente.</li>
            <li>Revise os detalhes e realize a assinatura digital.</li>
        </ol>

        <p style="color: #666; font-size: 12px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 10px;">
            Este é um e-mail automático enviado pelo Sistema de Avaliação Docente SENAI Morvan Figueiredo.
        </p>
    </div>
    """
    return send_resend_email(teacher_email, subject, html_content, attachments=attachments)

def send_scheduling_notification_resend(scheduled):
    """Send notification email when a new schedule is created"""
    if not scheduled.teacher or not scheduled.teacher.user or not scheduled.teacher.user.email:
        return False
    
    teacher_email = scheduled.teacher.user.email
    teacher_user = scheduled.teacher.user
    subject = f"Novo Agendamento de Avaliação - {scheduled.teacher.name}"
    
    # Credenciais e Link
    system_link = os.environ.get("DOMAIN") or "https://web-production-a43ed.up.railway.app/"
    username = teacher_user.username
    
    # Senha - Tenta pegar a senha real de forma mais abrangente
    password_info = None
    
    # PRIORIDADE: Seguir a mesma lógica do perfil para consistência total
    # 1. Tenta o campo permanentemente armazenado no banco
    if hasattr(teacher_user, 'plain_password') and teacher_user.plain_password:
        password_info = teacher_user.plain_password
    
    # 2. Tenta atributos temporários no objeto user (Request atual)
    if not password_info:
        if hasattr(teacher_user, '_password_plain') and teacher_user._password_plain:
            password_info = teacher_user._password_plain
        elif hasattr(teacher_user, '_temp_password') and teacher_user._temp_password:
            password_info = teacher_user._temp_password
    
    # 3. Tenta buscar na sessão (Última ação do usuário)
    if not password_info:
        from flask import session
        recent_creds = session.get('new_teacher_credentials')
        if recent_creds and recent_creds.get('username') == teacher_user.username:
            password_info = recent_creds.get('password')

    # 4. Fallback final (Caso nada funcione)
    if not password_info:
        password_info = teacher_user.username # Login/NIF como padrão de sistema

    # Gerar PDF de credenciais
    attachments = []
    try:
        pdf_buffer = generate_teacher_credentials_pdf(
            scheduled.teacher.name, 
            scheduled.teacher.nif, 
            username, 
            password_info
        )
        import base64
        attachments.append({
            "filename": f"Credenciais_{username}.pdf",
            "content": base64.b64encode(pdf_buffer.getvalue()).decode('utf-8')
        })
    except Exception as pdf_err:
        current_app.logger.error(f"Erro ao gerar PDF de credenciais para anexo: {pdf_err}")

    html_content = f"""
    <div style="font-family: Arial, sans-serif; line-height: 1.6; color: #333;">
        <h1 style="color: #1976d2;">Novo Agendamento de Avaliação</h1>
        <p>Olá <strong>{scheduled.teacher.name}</strong>,</p>
        <p>Um novo agendamento de avaliação foi criado para você.</p>
        
        <div style="background-color: #f5f5f5; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <h2 style="margin-top: 0; font-size: 18px;">Detalhes do Agendamento:</h2>
            <ul style="list-style: none; padding-left: 0;">
                <li><strong>Mês/Ano:</strong> {scheduled.scheduled_month}/{scheduled.scheduled_year}</li>
                <li><strong>Unidade Curricular:</strong> {scheduled.curricular_unit.name if scheduled.curricular_unit else 'N/A'}</li>
            </ul>
        </div>

        <div style="background-color: #e3f2fd; padding: 15px; border-radius: 5px; margin: 20px 0;">
            <h2 style="margin-top: 0; font-size: 18px; color: #1976d2;">Acesso ao Sistema:</h2>
            <p>Suas credenciais de acesso (Login e Senha) estão no <strong>PDF em anexo</strong> a este e-mail.</p>
            <p style="margin-bottom: 5px;"><strong>Link do Sistema:</strong> <a href="{system_link}">{system_link}</a></p>
        </div>

        <p>Acesse o sistema para mais informações e para se preparar para a avaliação.</p>

        <p style="color: #666; font-size: 12px; margin-top: 30px; border-top: 1px solid #eee; padding-top: 10px;">
            Este é um e-mail automático enviado pelo Sistema de Avaliação Docente SENAI Morvan Figueiredo.
        </p>
    </div>
    """
    return send_resend_email(teacher_email, subject, html_content, attachments=attachments)
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT
from datetime import datetime
import pandas as pd
import sendgrid
from sendgrid.helpers.mail import Mail, Email, To, Content

def get_or_create_current_semester():
    """Get or create current semester based on current date"""
    from models import Semester
    from models import db
    
    current_date = datetime.now()
    current_year = current_date.year
    current_month = current_date.month
    
    # Determine semester (1st = Jan-Jul, 2nd = Aug-Dec)
    semester_number = 1 if current_month <= 7 else 2
    semester_name = f"{current_year}.{semester_number}"
    
    current_semester = Semester.query.filter_by(
        year=current_year, 
        number=semester_number
    ).first()
    
    if not current_semester:
        # Create semester automatically based on current date
        start_date = datetime(current_year, 1, 1) if semester_number == 1 else datetime(current_year, 8, 1)
        end_date = datetime(current_year, 7, 31) if semester_number == 1 else datetime(current_year, 12, 31)
        
        current_semester = Semester(
            name=semester_name,
            year=current_year,
            number=semester_number,
            start_date=start_date,
            end_date=end_date,
            is_active=True
        )
        db.session.add(current_semester)
        db.session.commit()
    
    return current_semester
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

def save_uploaded_file(file):
    """Save uploaded file and return file info"""
    if file and file.filename:
        # Generate unique filename
        filename = secure_filename(file.filename)
        unique_filename = str(uuid.uuid4()) + '_' + filename
        
        # Save file
        file_path = os.path.join(current_app.config['UPLOAD_FOLDER'], unique_filename)
        file.save(file_path)
        
        return {
            'filename': unique_filename,
            'original_filename': filename,
            'file_path': file_path,
            'file_size': os.path.getsize(file_path),
            'mime_type': file.mimetype
        }
    return None

def send_credentials_email(teacher_email, teacher_data, password):
    """Send credentials email to newly created teacher"""
    if not teacher_email:
        return False
        
    try:
        msg = Message(
            subject=f'Credenciais de Acesso - Sistema de Avaliação Docente SENAI',
            recipients=[teacher_email],
            sender=current_app.config.get('MAIL_DEFAULT_SENDER', 'noreply@senai.br')
        )
        
        msg.body = f"""
Prezado(a) {teacher_data['teacher_name']},

Bem-vindo(a) ao Sistema de Avaliação Docente do SENAI Morvan Figueiredo!

Suas credenciais de acesso foram criadas com sucesso:

DADOS PESSOAIS:
Nome: {teacher_data['teacher_name']}
NIF: {teacher_data['teacher_nif']}
Área: {teacher_data['teacher_area']}

CREDENCIAIS DE ACESSO:
Usuário (Login): {teacher_data['username']}
Senha: {password}

INSTRUÇÕES PARA PRIMEIRO ACESSO:
1. Acesse o sistema usando o usuário e senha fornecidos acima
2. Após o login, clique no seu nome no canto superior direito
3. Selecione "Alterar Senha" para definir uma nova senha segura
4. Guarde suas novas credenciais em local seguro

IMPORTANTE:
- Altere sua senha no primeiro acesso por motivos de segurança
- Mantenha suas credenciais em sigilo
- Em caso de dúvidas, entre em contato com a coordenação

Atenciosamente,
Coordenação Pedagógica
SENAI Morvan Figueiredo

---
Este é um email automático. Por favor, não responda.
"""
        
        # Send email via SMTP (working on paid Railway)
        try:
            # Force immediate send and log everything
            current_app.logger.info(f"🔄 Iniciando envio de credenciais para {teacher_email}")
            current_app.logger.info(f"🔧 SMTP Server: {current_app.config.get('MAIL_SERVER')}:{current_app.config.get('MAIL_PORT')}")
            
            mail.send(msg)
            
            current_app.logger.info(f"✅ Credentials email sent successfully via SMTP to {teacher_email}")
            return True
        except Exception as smtp_error:
            current_app.logger.error(f"❌ SMTP failed for credentials to {teacher_email}. Detailed error: {str(smtp_error)}")
            current_app.logger.error(f"🔍 Error type: {type(smtp_error).__name__}")
            
            # Log specific Gmail errors
            error_str = str(smtp_error).lower()
            if '534' in error_str or 'password' in error_str:
                current_app.logger.error("🚨 Gmail authentication error - App password needed")
            elif '535' in error_str:
                current_app.logger.error("🚨 Gmail credentials invalid")
            elif 'connection' in error_str and 'refused' in error_str:
                current_app.logger.error("🚨 SMTP connection refused - possible Railway blocking")
            
            return False
        
    except Exception as e:
        current_app.logger.error(f"Error sending credentials email to {teacher_email}: {str(e)}")
        return False

def send_evaluation_email(teacher_email, evaluation, teacher_user=None, report_path=None):
    """Send evaluation notification email"""
    if not teacher_email:
        return False
    
    # Check if email is properly configured
    mail_server = current_app.config.get('MAIL_SERVER', 'localhost')
    mail_username = current_app.config.get('MAIL_USERNAME', '')
    
    # Skip email if not properly configured (avoid hanging on localhost)
    if mail_server == 'localhost' and not mail_username:
        current_app.logger.warning("Email not configured properly - skipping email sending")
        return False
    
    # Additional check for valid email server
    if not mail_server or mail_server.strip() == '':
        current_app.logger.warning("No email server configured - skipping email sending")
        return False
        
    try:
        # Get teacher user credentials if linked  
        teacher_credentials = ""
        if teacher_user and hasattr(teacher_user, '_password_plain'):
            # Show actual password if available (from recent registration)
            teacher_credentials = f"""
ACESSO AO SISTEMA:
Para acessar o sistema e assinar sua avaliação, utilize:
- Usuário: {teacher_user.username}
- Senha: {teacher_user._password_plain}

IMPORTANTE: Altere sua senha no primeiro acesso por motivos de segurança.
Guarde essas credenciais em local seguro.
"""
        elif teacher_user:
            # Include username, password will be shown generically for security
            teacher_credentials = f"""
ACESSO AO SISTEMA:
Para acessar o sistema e assinar sua avaliação, utilize:
- Usuário: {teacher_user.username}
- Senha: Sua senha é a mesma fornecida pela coordenação

IMPORTANTE: Altere sua senha no primeiro acesso por motivos de segurança.
Se você não lembra da senha, entre em contato com a coordenação.
"""
        elif evaluation.teacher.user:
            teacher_credentials = f"""
ACESSO AO SISTEMA:
Para acessar o sistema e assinar sua avaliação, utilize:
- Usuário: {evaluation.teacher.user.username}
- Senha: Sua senha é a mesma fornecida pela coordenação

IMPORTANTE: Altere sua senha no primeiro acesso por motivos de segurança.
Se você não lembra da senha, entre em contato com a coordenação.
"""
        else:
            teacher_credentials = """
ACESSO AO SISTEMA:
Entre em contato com a coordenação para obter suas credenciais de acesso ao sistema.
Seu usuário será seu número SN (NIF) e a senha será fornecida pela coordenação.
"""
        
        msg = Message(
            subject=f'Nova Avaliação Docente - Assinatura Necessária - {evaluation.evaluation_date.strftime("%d/%m/%Y")}',
            recipients=[teacher_email],
            sender=current_app.config.get('MAIL_DEFAULT_SENDER', 'noreply@senai.br')
        )
        
        msg.body = f"""
Prezado(a) {evaluation.teacher.name},

Você recebeu uma nova avaliação docente que precisa ser assinada no sistema.

DETALHES DA AVALIAÇÃO:
Curso: {evaluation.course.name}
Data: {evaluation.evaluation_date.strftime("%d/%m/%Y")}
Período: {evaluation.period}
Avaliador: {evaluation.evaluator.name if evaluation.evaluator else 'Coordenação'}

RESULTADOS:
Planejamento: {evaluation.calculate_planning_percentage()}% atendido
Condução da aula: {evaluation.calculate_class_percentage()}% atendido

{teacher_credentials}

PRÓXIMOS PASSOS:
1. Acesse o sistema usando suas credenciais
2. Vá até a seção "Avaliações"
3. Localize sua avaliação do dia {evaluation.evaluation_date.strftime("%d/%m/%Y")}
4. Revise os dados e assine digitalmente

Observações gerais:
{evaluation.general_observations or 'Nenhuma observação adicional registrada.'}

IMPORTANTE: Esta avaliação precisa ser assinada no sistema para ser finalizada.

Atenciosamente,
Coordenação Pedagógica
SENAI Morvan Figueiredo

---
Este é um e-mail automático do Sistema de Avaliação Docente.
Para dúvidas, entre em contato com a coordenação.
"""
        
        if report_path and os.path.exists(report_path):
            with open(report_path, 'rb') as f:
                msg.attach(f"relatorio_{evaluation.teacher.name.replace(' ', '_')}.pdf", 
                          "application/pdf", f.read())
        
        # Send email with SMTP timeout (safer than signals)
        try:
            mail.send(msg)
            current_app.logger.info(f"Email successfully sent to {teacher_email}")
            return True
        except Exception as smtp_error:
            current_app.logger.error(f"SMTP error sending email to {teacher_email}: {str(smtp_error)}")
            return False
            
    except Exception as e:
        current_app.logger.error(f"Error sending email to {teacher_email}: {str(e)}")
        return False

def send_email_via_sendgrid(to_email, subject, content_text, sender_email=None):
    """Send email using SendGrid API (works on Railway free plan)"""
    sendgrid_api_key = current_app.config.get('SENDGRID_API_KEY')
    
    if not sendgrid_api_key:
        current_app.logger.warning("SendGrid API key not configured - skipping email")
        return False
        
    if not sender_email:
        sender_email = current_app.config.get('MAIL_DEFAULT_SENDER', 'noreply@senai.br')
    
    try:
        sg = sendgrid.SendGridAPIClient(api_key=sendgrid_api_key)
        
        from_email = Email(sender_email)
        to_email_obj = To(to_email)
        content = Content("text/plain", content_text)
        
        mail = Mail(from_email, to_email_obj, subject, content)
        
        response = sg.send(mail)
        
        if response.status_code in [200, 201, 202]:
            current_app.logger.info(f"Email sent successfully via SendGrid to {to_email} (status: {response.status_code})")
            return True
        else:
            current_app.logger.error(f"SendGrid error: Status {response.status_code}, Body: {response.body}")
            return False
            
    except Exception as e:
        current_app.logger.error(f"SendGrid error sending email to {to_email}: {str(e)}")
        return False

def send_simple_evaluation_email(teacher_email, email_data):
    """Send evaluation notification email with primitive data (thread-safe)"""
    if not teacher_email:
        return False
        
    try:
        # Build credentials section
        teacher_credentials = ""
        if email_data.get('teacher_username'):
            teacher_credentials = f"""
ACESSO AO SISTEMA:
Para acessar o sistema e assinar sua avaliação, utilize:
- Usuário: {email_data['teacher_username']}
- Senha: Sua senha é a mesma fornecida pela coordenação

IMPORTANTE: Altere sua senha no primeiro acesso por motivos de segurança.
Se você não lembra da senha, entre em contato com a coordenação.
"""
        else:
            teacher_credentials = """
ACESSO AO SISTEMA:
Entre em contato com a coordenação para obter suas credenciais de acesso ao sistema.
Seu usuário será seu número SN (NIF) e a senha será fornecida pela coordenação.
"""
        
        msg = Message(
            subject=f'Nova Avaliação Docente - Assinatura Necessária - {email_data["evaluation_date"]}',
            recipients=[teacher_email],
            sender=current_app.config.get('MAIL_DEFAULT_SENDER', 'noreply@senai.br')
        )
        
        msg.body = f"""
Prezado(a) {email_data['teacher_name']},

Você recebeu uma nova avaliação docente que precisa ser assinada no sistema.

DETALHES DA AVALIAÇÃO:
Curso: {email_data['course_name']}
Data: {email_data['evaluation_date']}
Período: {email_data['period']}
Avaliador: {email_data['evaluator_name']}

RESULTADOS:
Planejamento de Aula: {email_data['planning_percentage']:.1f}%
Condução da Aula: {email_data['class_percentage']:.1f}%

{teacher_credentials}

PRÓXIMOS PASSOS:
1. Acesse o sistema de avaliação docente
2. Localize sua avaliação na área "Minhas Avaliações"
3. Revise os dados da avaliação
4. Assine digitalmente para confirmar ciência

IMPORTANTE:
- Você tem 7 dias para assinar esta avaliação
- Após a assinatura, você receberá uma cópia em PDF por email
- Em caso de dúvidas, entre em contato com a coordenação

Atenciosamente,
Coordenação Pedagógica
SENAI Morvan Figueiredo

---
Este é um email automático. Por favor, não responda.
"""
        
        # Send email via SMTP (working on paid Railway)
        try:
            # Force immediate send and log everything
            current_app.logger.info(f"🔄 Iniciando envio de avaliação para {teacher_email}")
            current_app.logger.info(f"🔧 SMTP Server: {current_app.config.get('MAIL_SERVER')}:{current_app.config.get('MAIL_PORT')}")
            
            mail.send(msg)
            
            current_app.logger.info(f"✅ Evaluation email sent successfully via SMTP to {teacher_email}")
            return True
        except Exception as smtp_error:
            current_app.logger.error(f"❌ SMTP failed for evaluation to {teacher_email}. Detailed error: {str(smtp_error)}")
            current_app.logger.error(f"🔍 Error type: {type(smtp_error).__name__}")
            
            # Log specific Gmail errors
            error_str = str(smtp_error).lower()
            if '534' in error_str or 'password' in error_str:
                current_app.logger.error("🚨 Gmail authentication error - App password needed")
            elif '535' in error_str:
                current_app.logger.error("🚨 Gmail credentials invalid") 
            elif 'connection' in error_str and 'refused' in error_str:
                current_app.logger.error("🚨 SMTP connection refused - possible Railway blocking")
                
            return False
            
    except Exception as e:
        current_app.logger.error(f"Error sending simple evaluation email to {teacher_email}: {str(e)}")
        return False

def generate_evaluation_report(evaluation):
    """Generate PDF report for a single evaluation"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], 
                                fontSize=16, spaceAfter=30, alignment=1)
    story.append(Paragraph("RELATÓRIO DE ACOMPANHAMENTO DOCENTE", title_style))
    story.append(Spacer(1, 20))
    
    # Custom style for checklist items to ensure wrapping
    criteria_style = ParagraphStyle('CriteriaStyle', 
                                  parent=styles['Normal'], 
                                  fontSize=9, 
                                  leading=11, 
                                  wordWrap='LTR')
    
    # Basic Info Table
    basic_info = [
        ['Docente:', evaluation.teacher.name],
        ['Curso:', evaluation.course.name],
        ['Período:', evaluation.period],
        ['Data:', evaluation.evaluation_date.strftime("%d/%m/%Y")]
    ]
    
    basic_table = Table(basic_info, colWidths=[2*inch, 4*inch])
    basic_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(basic_table)
    story.append(Spacer(1, 20))
    
    # Results Summary
    planning_pct = evaluation.calculate_planning_percentage()
    class_pct = evaluation.calculate_class_percentage()
    
    results = [
        ['RESUMO DOS RESULTADOS', ''],
        ['Planejamento:', f'{planning_pct}% atendido'],
        ['Condução da aula:', f'{class_pct}% atendido'],
    ]
    
    results_table = Table(results, colWidths=[2*inch, 4*inch])
    results_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 1), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(results_table)
    story.append(Spacer(1, 20))
    
    # Detailed Evaluation Criteria - Planning
    story.append(Paragraph("<b>CRITÉRIOS DE AVALIAÇÃO - PLANEJAMENTO</b>", styles['Heading3']))
    
    planning_criteria = [['Critério', 'Resposta']]
    planning_items = [item for item in evaluation.checklist_items if item.category == 'planning']
    
    # Sort items - relationship already has order_by but explicit sort is safer
    planning_items.sort(key=lambda x: x.display_order if x.display_order is not None else 0)
    
    for item in planning_items:
        planning_criteria.append([
            Paragraph(item.label, criteria_style), 
            Paragraph(item.value or 'N/A', criteria_style)
        ])
    
    # Fallback to hardcoded criteria if no dynamic items exist (for very old evaluations)
    if not planning_items:
        planning_criteria = [
            ['Critério', 'Resposta'],
            [Paragraph('Elabora cronograma de aula', styles['Normal']), Paragraph(evaluation.planning_schedule or 'N/A', styles['Normal'])],
            [Paragraph('Planeja a aula', styles['Normal']), Paragraph(evaluation.planning_lesson_plan or 'N/A', styles['Normal'])],
            [Paragraph('Planeja instrumentos de avaliação', styles['Normal']), Paragraph(evaluation.planning_evaluation or 'N/A', styles['Normal'])],
            [Paragraph('Conhece documentos estruturantes', styles['Normal']), Paragraph(evaluation.planning_documents or 'N/A', styles['Normal'])],
            [Paragraph('Utiliza instrumentos diversificados', styles['Normal']), Paragraph(evaluation.planning_diversified or 'N/A', styles['Normal'])],
            [Paragraph('Prepara previamente o local', styles['Normal']), Paragraph(evaluation.planning_local_work or 'N/A', styles['Normal'])],
            [Paragraph('Disponibiliza ferramentas', styles['Normal']), Paragraph(evaluation.planning_tools or 'N/A', styles['Normal'])],
            [Paragraph('Portal Educacional', criteria_style), Paragraph(evaluation.planning_educational_portal or 'N/A', criteria_style)],
        ]
    
    planning_table = Table(planning_criteria, colWidths=[4.2*inch, 1.8*inch], splitByRow=1)
    planning_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(planning_table)
    story.append(Spacer(1, 20))
    
    # Detailed Evaluation Criteria - Classroom
    story.append(Paragraph("<b>CRITÉRIOS DE AVALIAÇÃO - CONDUÇÃO DA AULA</b>", styles['Heading3']))
    
    class_criteria = [['Critério', 'Resposta']]
    class_items = [item for item in evaluation.checklist_items if item.category == 'class']
    class_items.sort(key=lambda x: x.display_order if x.display_order is not None else 0)
    
    for item in class_items:
        class_criteria.append([
            Paragraph(item.label, criteria_style), 
            Paragraph(item.value or 'N/A', criteria_style)
        ])
    
    # Fallback to hardcoded criteria if no dynamic items exist
    if not class_items:
        class_criteria = [
            ['Critério', 'Resposta'],
            [Paragraph('Apresentação pessoal', styles['Normal']), Paragraph(evaluation.class_presentation or 'N/A', styles['Normal'])],
            [Paragraph('Conhecimento dos assuntos', styles['Normal']), Paragraph(evaluation.class_knowledge or 'N/A', styles['Normal'])],
            [Paragraph('Acompanha desempenho', styles['Normal']), Paragraph(evaluation.class_student_performance or 'N/A', styles['Normal'])],
            [Paragraph('Registra ocorrências', styles['Normal']), Paragraph(evaluation.class_attendance or 'N/A', styles['Normal'])],
            [Paragraph('Realiza levantamento de dificuldades', styles['Normal']), Paragraph(evaluation.class_difficulties or 'N/A', styles['Normal'])],
            [Paragraph('Aprendizado teórico e prático', styles['Normal']), Paragraph(evaluation.class_theoretical_practical or 'N/A', styles['Normal'])],
            [Paragraph('Retoma aula anterior', styles['Normal']), Paragraph(evaluation.class_previous_lesson or 'N/A', styles['Normal'])],
            [Paragraph('Explicita objetivos', styles['Normal']), Paragraph(evaluation.class_objectives or 'N/A', styles['Normal'])],
            [Paragraph('Propõe questões', styles['Normal']), Paragraph(evaluation.class_questions or 'N/A', styles['Normal'])],
            [Paragraph('Verifica assimilação', styles['Normal']), Paragraph(evaluation.class_content_assimilation or 'N/A', styles['Normal'])],
            [Paragraph('Estimula participação', styles['Normal']), Paragraph(evaluation.class_student_participation or 'N/A', styles['Normal'])],
            [Paragraph('Processo de recuperação', styles['Normal']), Paragraph(evaluation.class_recovery_process or 'N/A', styles['Normal'])],
            [Paragraph('Exercícios para estimular', styles['Normal']), Paragraph(evaluation.class_learning_exercises or 'N/A', styles['Normal'])],
            [Paragraph('Mantém disciplina', styles['Normal']), Paragraph(evaluation.class_discipline or 'N/A', styles['Normal'])],
            [Paragraph('Estratégias de ensino', styles['Normal']), Paragraph(evaluation.class_teaching_strategies or 'N/A', styles['Normal'])],
            [Paragraph('Orienta utilização de equipamentos', styles['Normal']), Paragraph(evaluation.class_machines_equipment or 'N/A', styles['Normal'])],
            [Paragraph('Procedimentos de segurança', criteria_style), Paragraph(evaluation.class_safety_procedures or 'N/A', criteria_style)],
        ]
    
    class_table = Table(class_criteria, colWidths=[4.2*inch, 1.8*inch], splitByRow=1)
    class_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(class_table)
    story.append(Spacer(1, 20))

    # Observations
    if evaluation.planning_observations:
        story.append(Paragraph("<b>Observações - Planejamento:</b>", styles['Heading3']))
        story.append(Paragraph(evaluation.planning_observations, styles['Normal']))
        story.append(Spacer(1, 12))
    
    if evaluation.class_observations:
        story.append(Paragraph("<b>Observações - Período da Aula:</b>", styles['Heading3']))
        story.append(Paragraph(evaluation.class_observations, styles['Normal']))
        story.append(Spacer(1, 12))
    
    if evaluation.general_observations:
        story.append(Paragraph("<b>Observações Gerais:</b>", styles['Heading3']))
        story.append(Paragraph(evaluation.general_observations, styles['Normal']))
        story.append(Spacer(1, 12))
    
    # Signatures section
    story.append(Spacer(1, 30))
    story.append(Paragraph("<b>Assinaturas:</b>", styles['Heading3']))
    
    signature_data = [
        ['Avaliado por:', evaluation.evaluator.name if evaluation.evaluator else 'N/A'],
        ['Data da avaliação:', evaluation.evaluation_date.strftime('%d/%m/%Y')],
        ['', ''],
        ['Docente:', evaluation.teacher.name],
        ['Data da assinatura:', evaluation.teacher_signature_date.strftime('%d/%m/%Y %H:%M') if evaluation.teacher_signature_date else 'Pendente']
    ]
    
    signature_table = Table(signature_data, colWidths=[1.5*inch, 4.5*inch])
    signature_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (0, -1), 'Helvetica-Bold'),
        ('FONTNAME', (1, 0), (1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, 1), 1, colors.black),
        ('GRID', (0, 3), (-1, 4), 1, colors.black)
    ]))
    
    story.append(signature_table)
    
    # Add signature images if available
    teacher_signature = evaluation.signatures and next((s for s in evaluation.signatures if s.signature_type == 'teacher'), None)
    if teacher_signature and teacher_signature.signature_data:
        story.append(Spacer(1, 10))
        story.append(Paragraph("<b>Assinatura Digital do Docente:</b>", styles['Normal']))
        try:
            # Decode base64 signature
            import base64
            from reportlab.lib.utils import ImageReader
            
            signature_b64 = teacher_signature.signature_data.split(',')[1] if ',' in teacher_signature.signature_data else teacher_signature.signature_data
            signature_bytes = base64.b64decode(signature_b64)
            signature_image = ImageReader(BytesIO(signature_bytes))
            
            # Create a table with signature image and border for better visibility
            from reportlab.platypus import Image
            signature_img = Image(signature_image, width=200, height=80)
            
            signature_img_table = Table([[signature_img]], colWidths=[220])
            signature_img_table.setStyle(TableStyle([
                ('ALIGN', (0, 0), (0, 0), 'LEFT'),
                ('VALIGN', (0, 0), (0, 0), 'MIDDLE'),
                ('BACKGROUND', (0, 0), (0, 0), colors.white),
                ('GRID', (0, 0), (0, 0), 2, colors.black),
                ('TOPPADDING', (0, 0), (0, 0), 8),
                ('BOTTOMPADDING', (0, 0), (0, 0), 8),
                ('LEFTPADDING', (0, 0), (0, 0), 10),
                ('RIGHTPADDING', (0, 0), (0, 0), 10)
            ]))
            story.append(signature_img_table)
        except Exception as e:
            story.append(Paragraph(f"[Assinatura digital registrada em {teacher_signature.signed_at.strftime('%d/%m/%Y %H:%M')}]", styles['Normal']))
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph(f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", 
                          styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_consolidated_report(teacher_id, start_date=None, end_date=None):
    """Generate consolidated report for a teacher"""
    from models import Evaluation, Teacher
    
    teacher = Teacher.query.get(teacher_id)
    if not teacher:
        return None
    
    query = Evaluation.query.filter_by(teacher_id=teacher_id)
    if start_date:
        query = query.filter(Evaluation.evaluation_date >= start_date)
    if end_date:
        query = query.filter(Evaluation.evaluation_date <= end_date)
    
    evaluations = query.order_by(Evaluation.evaluation_date.desc()).all()
    
    if not evaluations:
        return None
    
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []
    
    # Title
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], 
                                fontSize=16, spaceAfter=30, alignment=1)
    story.append(Paragraph(f"RELATÓRIO CONSOLIDADO - {teacher.name.upper()}", title_style))
    story.append(Spacer(1, 20))
    
    # Teacher info
    teacher_info = [
        ['Docente:', teacher.name],
        ['Área:', teacher.area],
        ['Disciplinas:', teacher.subjects or 'Não informado'],
        ['Total de Acompanhamentos:', str(len(evaluations))]
    ]
    
    teacher_table = Table(teacher_info, colWidths=[2*inch, 4*inch])
    teacher_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 0), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 0), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(teacher_table)
    story.append(Spacer(1, 20))
    
    # Evolution table
    evolution_data = [['Data', 'Curso', 'Planejamento', 'Condução da Aula']]
    
    for eval in evaluations:
        evolution_data.append([
            eval.evaluation_date.strftime("%d/%m/%Y"),
            eval.course.name,
            f"{eval.calculate_planning_percentage()}%",
            f"{eval.calculate_class_percentage()}%"
        ])
    
    evolution_table = Table(evolution_data, colWidths=[1.5*inch, 2*inch, 1.2*inch, 1.3*inch])
    evolution_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey])
    ]))
    
    story.append(Paragraph("<b>Evolução dos Acompanhamentos:</b>", styles['Heading3']))
    story.append(evolution_table)
    story.append(Spacer(1, 20))
    
    # Average performance
    avg_planning = sum(eval.calculate_planning_percentage() for eval in evaluations) / len(evaluations)
    avg_class = sum(eval.calculate_class_percentage() for eval in evaluations) / len(evaluations)
    
    avg_data = [
        ['DESEMPENHO MÉDIO', ''],
        ['Planejamento:', f'{avg_planning:.1f}%'],
        ['Condução da aula:', f'{avg_class:.1f}%']
    ]
    
    avg_table = Table(avg_data, colWidths=[2*inch, 4*inch])
    avg_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('BACKGROUND', (0, 1), (0, -1), colors.grey),
        ('TEXTCOLOR', (0, 1), (0, -1), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
        ('BACKGROUND', (1, 1), (1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black)
    ]))
    
    story.append(avg_table)
    
    # Footer
    story.append(Spacer(1, 30))
    story.append(Paragraph(f"Relatório gerado em {datetime.now().strftime('%d/%m/%Y às %H:%M')}", 
                          styles['Normal']))
    
    doc.build(story)
    buffer.seek(0)
    return buffer

def generate_teachers_excel_template():
    """Generate Excel template for teacher import"""
    buffer = BytesIO()
    
    # Create sample data with headers
    data = {
        'NIF': ['SN1234567', 'SN7654321'],
        'Nome': ['João Silva', 'Maria Santos'],
        'Área': ['Eletrônica', 'Informática']
    }
    
    df = pd.DataFrame(data)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Docentes"
    
    # Add header row
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instruções")
    instructions = [
        ["INSTRUÇÕES PARA IMPORTAÇÃO DE DOCENTES"],
        [""],
        ["1. Preencha os dados dos docentes na aba 'Docentes'"],
        ["2. Campos obrigatórios: NIF, Nome, Área"],
        ["3. NIF: número de identificação funcional no formato SN1234567"],
        ["4. Nome: nome completo do docente"],
        ["5. Área: área de atuação do docente (ex: Eletrônica, Informática)"],
        [""],
        ["IMPORTANTE:"],
        ["- Não altere os nomes das colunas"],
        ["- Mantenha o formato Excel (.xlsx)"],
        ["- Remova as linhas de exemplo antes de importar seus dados"],
        ["- O formato simplificado permite avaliação rápida por curso e área"],
    ]
    
    for row in instructions:
        instructions_ws.append(row)
    
    # Style instructions
    instructions_ws['A1'].font = Font(bold=True, size=14)
    instructions_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    instructions_ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    
    for i in range(11, 16):  # Important section
        instructions_ws[f'A{i}'].font = Font(bold=True)
    
    # Auto-adjust column width for instructions
    for column in instructions_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        instructions_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def process_teachers_excel_import(file_path):
    """Process Excel file and import teachers"""
    from models import Teacher, Course
    from models import db
    import sys
    
    # Ensure UTF-8 encoding
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    
    try:
        # Ensure the file is read with proper encoding
        import locale
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
        
        # Read Excel file with explicit engine
        df = pd.read_excel(file_path, sheet_name='Docentes', engine='openpyxl')
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        results = {
            'success': 0,
            'errors': [],
            'warnings': []
        }
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get('NIF', '')) or str(row.get('NIF', '')).strip() == '':
                    continue
                
                # Extract data with defaults - proper string handling
                nif = str(row['NIF']).strip().upper() if not pd.isna(row['NIF']) else ''
                name = str(row.get('Nome', '')).strip() if not pd.isna(row.get('Nome', '')) else ''
                area = str(row.get('Área', '')).strip() if not pd.isna(row.get('Área', '')) else ''
                
                # Validate required fields
                if not nif:
                    results['errors'].append(f'Linha {index + 2}: NIF é obrigatório')
                    continue
                    
                if not nif.startswith('SN') or len(nif) != 9:
                    results['errors'].append(f'Linha {index + 2}: NIF deve estar no formato SN1234567')
                    continue
                
                if not name:
                    results['errors'].append(f'Linha {index + 2}: Nome é obrigatório')
                    continue
                    
                if not area:
                    results['errors'].append(f'Linha {index + 2}: Área é obrigatória')
                    continue
                
                # Check if teacher already exists
                existing_teacher = Teacher.query.filter_by(nif=nif).first()
                if existing_teacher:
                    results['warnings'].append(f'Linha {index + 2}: Docente com NIF "{nif}" já existe, pulando...')
                    continue
                
                # Create new teacher
                teacher = Teacher()  # type: ignore
                teacher.nif = nif
                teacher.name = name
                teacher.area = area
                
                db.session.add(teacher)
                db.session.flush()  # Get teacher ID before commit
                
                # Create user account for teacher
                from models import User
                import random
                import string
                
                # Generate random password
                password = ''.join(random.choices(string.ascii_letters + string.digits, k=8))
                
                user = User()  # type: ignore
                user.username = nif.lower()
                user.name = name
                user.role = 'teacher'
                user.set_password(password)
                
                db.session.add(user)
                db.session.flush()
                
                teacher.user_id = user.id
                
                # Store the password for later notification (you might want to email this)
                results['warnings'].append(f'Docente {name} (NIF: {nif}) - Senha gerada: {password}')
                
                results['success'] += 1
                
            except Exception as e:
                db.session.rollback()  # Rollback the failed transaction
                results['errors'].append(f'Linha {index + 2}: Erro ao processar - {str(e)}')
        
        # Commit all changes
        if results['success'] > 0:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                results['errors'].append(f'Erro ao salvar dados: {str(e)}')
                results['success'] = 0
        
        return results
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': 0,
            'errors': [f'Erro ao processar arquivo Excel: {str(e)}'],
            'warnings': []
        }

def generate_courses_excel_template():
    """Generate Excel template for course import"""
    buffer = BytesIO()
    
    # Create sample data with headers - now supporting up to 10 curricular units
    data = {
        'Nome do Curso': ['Técnico em Eletrônica', 'Técnico em Informática'],
        'Unidade Curricular 1': ['Eletrônica Digital', 'Programação Web'],
        'Unidade Curricular 2': ['Circuitos Elétricos', 'Banco de Dados'],
        'Unidade Curricular 3': ['Microcontroladores', 'Redes de Computadores'],
        'Unidade Curricular 4': ['Automação Industrial', 'Desenvolvimento Mobile'],
        'Unidade Curricular 5': ['', 'Segurança da Informação'],
        'Unidade Curricular 6': ['', ''],
        'Unidade Curricular 7': ['', ''],
        'Unidade Curricular 8': ['', ''],
        'Unidade Curricular 9': ['', ''],
        'Unidade Curricular 10': ['', '']
    }
    
    df = pd.DataFrame(data)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cursos"
    
    # Add header row
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Add instructions sheet
    instructions_ws = wb.create_sheet("Instruções")
    instructions = [
        ["INSTRUÇÕES PARA IMPORTAÇÃO DE CURSOS"],
        [""],
        ["1. Preencha os dados dos cursos na aba 'Cursos'"],
        ["2. Campos obrigatórios: Nome do Curso, pelo menos 1 Unidade Curricular"],
        ["3. Nome do Curso: nome completo do curso (ex: Técnico em Eletrônica)"],
        ["4. Unidades Curriculares 1-10: disciplinas do curso"],
        ["5. Preencha apenas as unidades curriculares que existem"],
        ["6. Deixe em branco as unidades curriculares não utilizadas"],
        [""],
        ["IMPORTANTE:"],
        ["- Não altere os nomes das colunas"],
        ["- Mantenha o formato Excel (.xlsx)"],
        ["- Remova as linhas de exemplo antes de importar seus dados"],
        ["- Cada curso pode ter até 10 unidades curriculares"],
        ["- Isso facilita a avaliação de docentes por unidade curricular específica"],
    ]
    
    for row in instructions:
        instructions_ws.append(row)
    
    # Style instructions
    instructions_ws['A1'].font = Font(bold=True, size=14)
    instructions_ws['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    instructions_ws['A1'].font = Font(color="FFFFFF", bold=True, size=14)
    
    for i in range(10, 14):  # Important section
        instructions_ws[f'A{i}'].font = Font(bold=True)
    
    # Auto-adjust column width for instructions
    for column in instructions_ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 80)
        instructions_ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def process_courses_excel_import(file_path):
    """Process Excel file and import courses with curricular units"""
    from models import Course, CurricularUnit
    from models import db
    import sys
    
    # Ensure UTF-8 encoding
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    
    try:
        # Ensure the file is read with proper encoding
        import locale
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
        
        # Read Excel file with explicit engine
        df = pd.read_excel(file_path, sheet_name='Cursos', engine='openpyxl')
        
        # Clean column names
        df.columns = df.columns.str.strip()
        
        results = {
            'success': 0,
            'errors': [],
            'warnings': []
        }
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get('Nome do Curso', '')) or str(row.get('Nome do Curso', '')).strip() == '':
                    continue
                
                # Extract course name - proper string handling
                course_name = str(row['Nome do Curso']).strip() if not pd.isna(row['Nome do Curso']) else ''
                
                # Validate required fields
                if not course_name:
                    results['errors'].append(f'Linha {index + 2}: Nome do Curso é obrigatório')
                    continue
                
                # Extract curricular units (up to 10)
                curricular_units = []
                for i in range(1, 11):
                    unit_column = f'Unidade Curricular {i}'
                    if unit_column in row and not pd.isna(row.get(unit_column)):
                        unit_name = str(row[unit_column]).strip()
                        if unit_name:
                            curricular_units.append(unit_name)
                
                # Note: Curricular units are now optional
                
                # Check if course already exists
                existing_course = Course.query.filter_by(name=course_name).first()
                if existing_course:
                    results['warnings'].append(f'Linha {index + 2}: Curso "{course_name}" já existe, atualizando unidades curriculares...')
                    course = existing_course
                else:
                    # Create new course
                    from datetime import datetime
                    current_year = datetime.now().year
                    course = Course()  # type: ignore
                    course.name = course_name
                    course.period = f"1° Sem/{current_year}"  # Default period based on current year
                    course.curriculum_component = "Múltiplas Unidades Curriculares"
                    course.class_code = None
                    db.session.add(course)
                    db.session.flush()  # Get course ID
                
                # Add curricular units
                units_added = 0
                for unit_name in curricular_units:
                    # Check if unit already exists for this course
                    existing_unit = CurricularUnit.query.filter_by(
                        name=unit_name, 
                        course_id=course.id
                    ).first()
                    
                    if not existing_unit:
                        unit = CurricularUnit()  # type: ignore
                        unit.name = unit_name
                        unit.course_id = course.id
                        unit.code = None
                        unit.workload = None
                        unit.description = f"Unidade curricular do curso {course_name}"
                        unit.is_active = True
                        db.session.add(unit)
                        units_added += 1
                
                if not existing_course:
                    results['success'] += 1
                
                if units_added > 0:
                    results['warnings'].append(f'Linha {index + 2}: {units_added} unidades curriculares adicionadas ao curso "{course_name}"')
                
            except Exception as e:
                db.session.rollback()  # Rollback the failed transaction
                results['errors'].append(f'Linha {index + 2}: Erro ao processar - {str(e)}')
        
        # Commit all changes
        if results['success'] > 0 or any('unidades curriculares adicionadas' in w for w in results['warnings']):
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                results['errors'].append(f'Erro ao salvar dados: {str(e)}')
                results['success'] = 0
        
        return results
        
    except Exception as e:
        db.session.rollback()
        return {
            'success': 0,
            'errors': [f'Erro ao processar arquivo Excel: {str(e)}'],
            'warnings': []
        }

def generate_curricular_units_excel_template():
    """Generate Excel template for curricular units import"""
    from io import BytesIO
    import pandas as pd
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    buffer = BytesIO()
    
    # Create sample data with headers
    data = {
        "Nome": ["Eletrônica Digital I", "Programação Web I", "Circuitos Elétricos"],
        "Código": ["ELD001", "PWB001", "CEL001"],
        "Curso": ["Técnico em Eletrônica", "Técnico em Informática", "Técnico em Eletrônica"],
        "Carga Horária": [80, 60, 120],
        "Descrição": ["Fundamentos de eletrônica digital", "Desenvolvimento de páginas web", "Conceitos básicos de circuitos elétricos"]
    }
    
    df = pd.DataFrame(data)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Unidades Curriculares"
    
    # Add header row
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    
    # Style the header row
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    # Save to buffer
    wb.save(buffer)
    buffer.seek(0)
    return buffer

def process_curricular_units_excel_import(file_path):
    """Process Excel file and import curricular units"""
    import pandas as pd
    from models import CurricularUnit, Course
    from models import db
    import sys
    
    # Ensure UTF-8 encoding
    if sys.stdout.encoding != 'utf-8':
        sys.stdout.reconfigure(encoding='utf-8')
    
    try:
        # Ensure the file is read with proper encoding
        import locale
        locale.setlocale(locale.LC_ALL, 'C.UTF-8')
        
        # Read Excel file with explicit engine  
        df = pd.read_excel(file_path, sheet_name="Unidades Curriculares", engine='openpyxl')
        
        results = {
            "success": 0,
            "errors": [],
            "warnings": []
        }
        
        for index, row in df.iterrows():
            try:
                # Skip empty rows
                if pd.isna(row.get("Nome", "")) or str(row.get("Nome", "")).strip() == "":
                    continue
                
                # Extract data - proper string handling
                name = str(row["Nome"]).strip() if not pd.isna(row["Nome"]) else ""
                code = str(row.get("Código", "")).strip() if not pd.isna(row.get("Código")) else ""
                course_name = str(row.get("Curso", "")).strip() if not pd.isna(row.get("Curso")) else ""
                description = str(row.get("Descrição", "")).strip() if not pd.isna(row.get("Descrição")) else ""
                
                # Validate required fields
                if not name:
                    results["errors"].append(f"Linha {index + 2}: Nome é obrigatório")
                    continue
                
                if not course_name:
                    results["errors"].append(f"Linha {index + 2}: Curso é obrigatório")
                    continue
                
                # Find course
                course = Course.query.filter_by(name=course_name).first()
                if not course:
                    results["errors"].append(f"Linha {index + 2}: Curso \"{course_name}\" não encontrado")
                    continue
                
                # Check if curricular unit already exists
                existing_unit = CurricularUnit.query.filter_by(name=name, course_id=course.id).first()
                if existing_unit:
                    results["warnings"].append(f"Linha {index + 2}: Unidade curricular \"{name}\" já existe para o curso \"{course_name}\", pulando...")
                    continue
                
                # Create new curricular unit
                unit = CurricularUnit()  # type: ignore
                unit.name = name
                unit.code = code if code else None
                unit.course_id = course.id
                unit.description = description if description else None
                unit.is_active = True
                
                db.session.add(unit)
                results["success"] += 1
                
            except Exception as e:
                db.session.rollback()  # Rollback the failed transaction
                results["errors"].append(f"Linha {index + 2}: Erro ao processar - {str(e)}")
        
        # Commit all changes
        if results["success"] > 0:
            try:
                db.session.commit()
            except Exception as e:
                db.session.rollback()
                results["errors"].append(f"Erro ao salvar dados: {str(e)}")
                results["success"] = 0
        
        return results
        
    except Exception as e:
        db.session.rollback()
        return {
            "success": 0,
            "errors": [f"Erro ao processar arquivo Excel: {str(e)}"],
            "warnings": []
        }

def generate_teacher_credentials_pdf(teacher_name, teacher_nif, username, password):
    """Generate PDF with teacher credentials for download - Optimized for single page"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, 
                           topMargin=0.5*inch, bottomMargin=0.5*inch,
                           leftMargin=0.75*inch, rightMargin=0.75*inch)
    styles = getSampleStyleSheet()
    story = []
    
    # Title - More compact
    title_style = ParagraphStyle('CustomTitle', parent=styles['Title'], 
                                fontSize=16, spaceAfter=15, alignment=TA_CENTER,
                                textColor=colors.darkblue)
    story.append(Paragraph("CREDENCIAIS DE ACESSO", title_style))
    
    # Institution info - Combined with subtitle
    institution_style = ParagraphStyle('Institution', parent=styles['Normal'],
                                     fontSize=12, spaceAfter=20, alignment=TA_CENTER,
                                     textColor=colors.darkgrey)
    story.append(Paragraph("Sistema de Avaliação Docente - SENAI Morvan Figueiredo", institution_style))
    
    # Credentials Box - Compact version without empty rows
    credentials_data = [
        ['DADOS PARA ACESSO', ''],
        ['Nome Completo:', teacher_name],
        ['NIF:', teacher_nif],
        ['Usuário (Login):', username],
        ['Senha Temporária:', password]
    ]
    
    credentials_table = Table(credentials_data, colWidths=[2.3*inch, 3.7*inch])
    credentials_table.setStyle(TableStyle([
        # Header styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.darkblue),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 13),
        ('SPAN', (0, 0), (-1, 0)),  # Merge header cells
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        
        # Data labels styling
        ('BACKGROUND', (0, 1), (0, 4), colors.lightblue),
        ('FONTNAME', (0, 1), (0, 4), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 11),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        
        # Data values styling
        ('BACKGROUND', (1, 1), (1, 4), colors.beige),
        ('FONTNAME', (1, 4), (1, 4), 'Helvetica-Bold'),  # Password in bold
        ('TEXTCOLOR', (1, 4), (1, 4), colors.red),  # Password in red
        
        # Border and padding - Reduced padding
        ('GRID', (0, 0), (-1, 4), 1, colors.black),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 12),
        ('RIGHTPADDING', (0, 0), (-1, -1), 12),
    ]))
    
    story.append(credentials_table)
    story.append(Spacer(1, 20))
    
    # Instructions - More compact
    instructions_style = ParagraphStyle('Instructions', parent=styles['Normal'],
                                      fontSize=10, spaceAfter=5,
                                      leftIndent=0.2*inch, rightIndent=0.2*inch,
                                      leading=14)
    
    story.append(Paragraph("<b>INSTRUÇÕES IMPORTANTES:</b>", instructions_style))
    story.append(Spacer(1, 8))
    
    # Compact instructions in a single paragraph with bullet points
    instructions_text = """
    • Use o usuário e senha acima para acessar o sistema<br/>
    • No primeiro acesso, altere sua senha por uma segura<br/>
    • Mantenha suas credenciais em local seguro<br/>
    • Contate a coordenação em caso de dúvidas
    """
    
    story.append(Paragraph(instructions_text, instructions_style))
    story.append(Spacer(1, 15))
    
    # Warning box - More compact
    warning_style = ParagraphStyle('Warning', parent=styles['Normal'],
                                 fontSize=9, alignment=TA_CENTER,
                                 textColor=colors.red,
                                 backColor=colors.mistyrose,
                                 borderColor=colors.red,
                                 borderWidth=1,
                                 borderPadding=8,
                                 spaceAfter=15)
    
    story.append(Paragraph("<b>⚠️ ATENÇÃO: Esta senha é temporária e deve ser alterada no primeiro acesso!</b>", warning_style))
    
    # Footer - Compact
    footer_style = ParagraphStyle('Footer', parent=styles['Normal'],
                                fontSize=8, alignment=TA_CENTER,
                                textColor=colors.grey,
                                spaceAfter=5)
    
    current_date = datetime.now().strftime("%d/%m/%Y às %H:%M")
    story.append(Paragraph(f"Credenciais geradas em {current_date}", footer_style))
    story.append(Paragraph("Sistema de Avaliação Docente - SENAI Morvan Figueiredo", footer_style))
    
    # Build PDF
    doc.build(story)
    buffer.seek(0)
    return buffer

